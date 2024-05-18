# # # open_workbook.py
from openpyxl import load_workbook
# import xlsxwriter
import openpyxl
# # #
# my_wb = openpyxl.Workbook()
# my_sheet = my_wb.active
# my_sheet.title="Student_Data"
# c1 = my_sheet.cell(row = 1, column = 1)
# c1.value = "Roll_Number"
# c2 = my_sheet.cell(row= 1 , column = 2)
# c2.value = "Name"
# c3 = my_sheet['C1']
# c3.value = "Percentage"
# # B2 = column = 2 & row = 2.
# c4 = my_sheet['D1']
# c4.value = "Grade"
# # my_wb.save("D:\Projects.xlsx")
# data = (
#     (10,'Narsh',90,'A'),
#     (20,'Adarsh',60,'B'),
#     (30,'Sparsh',70,'B')
# )
# for i in data:
#     my_sheet.append(i)
# my_wb.save('D:\Projects.xlsx')

def open_workbook(path):
    workbook = load_workbook(filename=path)
    print(f"Worksheet names: {workbook.sheetnames}")
    sheet = workbook.active
    print(sheet)
    print(f"The title of the Worksheet is: {sheet.title}")
    # for i in range(1, sheet.max_column + 1,sheet.max_column+1):
    #     cell=sheet.cell(row=i,column=i+1)
    #     print(cell.value)
    for row in range(0, sheet.max_row):
        for col in sheet.iter_cols(1, sheet.max_column):
            print(col[row].value)
if __name__ == "__main__":
    open_workbook("D:\Projects.xlsx")